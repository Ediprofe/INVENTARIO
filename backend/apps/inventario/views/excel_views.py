"""
Views para importación y exportación de Excel.

Este módulo contiene las vistas API para:
- Importar ítems desde archivos Excel
- Exportar ítems a archivos Excel
- Descargar plantilla de importación
"""
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from datetime import datetime
import openpyxl

from apps.inventario.models import ItemInventario
from apps.inventario.filters import ItemInventarioFilter
from apps.inventario.services import ExcelImportService, ImportValidationError
from apps.inventario.services.reset_import_service import ResetImportService
from apps.inventario.utils.excel_helpers import (
    apply_header_styles,
    adjust_column_widths,
    write_item_row,
    write_template_example_row,
    EXPORT_HEADERS,
    IMPORT_TEMPLATE_HEADERS
)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_items_excel(request):
    """
    Importa ítems desde archivo Excel según INICIAL.md líneas 980-1297.
    
    Implementa importación transaccional completa:
    - Si hay errores de validación, no se importa nada (todo o nada)
    - Auto-crea catálogos (Sedes, Ubicaciones, Responsables, Artículos)
    - Genera códigos automáticamente
    - Registra movimientos de inventario
    
    Request:
        POST con multipart/form-data
        Parámetro: file (archivo .xlsx o .xls)
    
    Response (éxito - 201):
        {
            "message": "Importación completada exitosamente",
            "created": 15,
            "errors": 0,
            "created_items": [{"row": 2, "codigo": "INV-001", "id": 1}],
            "error_details": [],
            "summary": {
                "sedes_creadas": 2,
                "ubicaciones_creadas": 5,
                "responsables_creados": 3,
                "articulos_creados": 10
            }
        }
    
    Response (error - 400):
        {
            "message": "Errores de validación encontrados",
            "created": 0,
            "errors": 3,
            "created_items": [],
            "error_details": [{"row": 3, "error": "Sede no puede estar vacía"}]
        }
    
    Columnas requeridas:
        Sede, Ubicacion, Articulo, Estado, Disponibilidad, Responsable
    
    Columnas opcionales:
        Placa, Marca, Serial, Descripcion, Observaciones
    """
    # Validar que se envió un archivo
    if 'file' not in request.FILES:
        return _error_response(
            'No se proporcionó ningún archivo',
            [{'row': 0, 'error': 'Archivo no encontrado'}]
        )
    
    excel_file = request.FILES['file']
    
    # Validar extensión
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return _error_response(
            'El archivo debe ser formato Excel (.xlsx o .xls)',
            [{'row': 0, 'error': 'Formato de archivo inválido'}]
        )
    
    # Procesar importación
    try:
        service = ExcelImportService()
        result = service.import_from_file(excel_file)
        return Response(result, status=status.HTTP_201_CREATED)
        
    except ImportValidationError as e:
        return _error_response('Errores de validación encontrados', e.errors)
        
    except ValueError as e:
        return _error_response(str(e), [{'row': 0, 'error': str(e)}])
        
    except Exception as e:
        return Response(
            {
                'message': f'Error al procesar archivo: {str(e)}',
                'created': 0,
                'errors': 1,
                'created_items': [],
                'error_details': [{'row': 0, 'error': str(e)}]
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_items_excel(request):
    """
    Exporta ítems a archivo Excel con filtros aplicados.
    
    Respeta todos los filtros del frontend para exportar solo
    los ítems visibles en la tabla actual.
    
    Query params opcionales:
        - sede, ubicacion, responsable, articulo (IDs)
        - estado, disponibilidad
        - search (búsqueda por código/nombre)
        - Ver ItemInventarioFilter para lista completa
    
    Response:
        Archivo Excel: inventario_YYYYMMDD_HHMMSS.xlsx
    """
    # Aplicar filtros
    queryset = ItemInventario.objects.select_related(
        'articulo', 'sede', 'ubicacion', 'responsable'
    )
    
    filterset = ItemInventarioFilter(request.GET, queryset=queryset)
    items = filterset.qs
    
    # Crear libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ítems Inventario"
    
    # Aplicar estilos y escribir datos
    apply_header_styles(ws, EXPORT_HEADERS)
    
    for row_num, item in enumerate(items, 2):
        write_item_row(ws, row_num, item)
    
    adjust_column_widths(ws)
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_template(request):
    """
    Descarga plantilla Excel para importación según INICIAL.md.
    
    La plantilla incluye:
    - Encabezados con columnas requeridas y opcionales
    - Fila de ejemplo con datos de muestra
    - Formato consistente con el proceso de importación
    
    Response:
        Archivo Excel: plantilla_importacion.xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"
    
    # Aplicar estilos y contenido
    apply_header_styles(ws, IMPORT_TEMPLATE_HEADERS)
    write_template_example_row(ws)
    
    # Ajustar anchos
    for col_num in range(1, len(IMPORT_TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_num)
        ].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="plantilla_importacion.xlsx"'
    )
    
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reset_import_excel(request):
    """
    Resetea el inventario completo e importa desde Excel multi-hoja.
    
    Esta vista ejecuta un proceso transaccional que:
    1. Elimina todos los ítems del inventario existente
    2. Crea/actualiza catálogos desde las hojas del Excel
    3. Importa los nuevos ítems
    
    El archivo debe contener 5 hojas:
    - Items: Datos del inventario principal
    - Sedes: Catálogo de sedes
    - Ubicaciones: Catálogo de ubicaciones
    - Articulos: Catálogo de artículos
    - Responsables: Catálogo de responsables
    
    Request:
        POST con multipart/form-data
        Parámetro: file (archivo .xlsx)
    
    Response (éxito - 200):
        {
            "success": true,
            "stats": {
                "items_eliminados": 100,
                "items_creados": 150,
                "sedes_creadas": 2,
                "ubicaciones_creadas": 10,
                "articulos_creados": 50,
                "responsables_creados": 20
            },
            "errors": []
        }
    
    Response (error - 400/500):
        {
            "success": false,
            "message": "Descripción del error",
            "errors": ["error1", "error2"]
        }
    """
    # Validar que se envió un archivo
    if 'file' not in request.FILES:
        return Response(
            {
                'success': False,
                'message': 'No se proporcionó ningún archivo',
                'errors': ['Archivo no encontrado']
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    excel_file = request.FILES['file']
    
    # Validar extensión
    if not excel_file.name.endswith('.xlsx'):
        return Response(
            {
                'success': False,
                'message': 'El archivo debe ser formato Excel (.xlsx)',
                'errors': ['Formato de archivo inválido. Solo se acepta .xlsx']
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Procesar reseteo e importación
    try:
        service = ResetImportService(excel_file)
        result = service.execute()
        
        return Response(result, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {
                'success': False,
                'message': f'Error al procesar archivo: {str(e)}',
                'errors': [str(e)]
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_reset_template(request):
    """
    Descarga plantilla Excel multi-hoja para reseteo e importación completa.
    
    La plantilla incluye 5 hojas con:
    - Encabezados descriptivos
    - 10 registros de ejemplo en cada hoja
    - Validaciones y buenas prácticas integradas
    
    Response:
        Archivo Excel: plantilla_reseteo_inventario.xlsx
    """
    from apps.inventario.utils.reset_template_generator import generate_reset_template
    
    try:
        wb = generate_reset_template()
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="plantilla_reseteo_inventario.xlsx"'
        )
        
        wb.save(response)
        return response
        
    except Exception as e:
        return Response(
            {
                'success': False,
                'message': f'Error al generar plantilla: {str(e)}',
                'errors': [str(e)]
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _error_response(message: str, errors: list) -> Response:
    """
    Construye una respuesta de error estándar.
    
    Args:
        message: Mensaje de error principal
        errors: Lista de errores detallados
        
    Returns:
        Response con estructura estándar de error
    """
    return Response(
        {
            'message': message,
            'created': 0,
            'errors': len(errors),
            'created_items': [],
            'error_details': errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )
