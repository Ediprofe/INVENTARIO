"""
Servicio para reseteo e importación masiva del inventario desde Excel.

Este servicio orquesta la eliminación completa del inventario existente
y la carga de un nuevo inventario desde un archivo Excel multi-hoja.

Características:
- Transaccional: Todo o nada (django.db.transaction.atomic)
- Auto-creación de catálogos (Sedes, Ubicaciones, Artículos, Responsables)
- Validaciones robustas en cada paso
- Generación automática de códigos
- Registro de historial de movimientos
"""

from typing import Dict, List, Any, Tuple
import logging
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from apps.inventario.models import (
    ItemInventario,
    Sede,
    Ubicacion,
    Articulo,
    Responsable,
    HistorialMovimiento,
)
from apps.inventario.models.choices import (
    TipoUbicacion,
    CategoriaArticulo,
    EstadoFisico,
    Disponibilidad,
    TipoDocumento,
    CargoResponsable,
    TipoMovimiento,
)

logger = logging.getLogger(__name__)

class ResetImportService:
    """
    Servicio para resetear e importar inventario completo desde Excel.
    """
    
    REQUIRED_SHEETS = ['Items', 'Sedes', 'Ubicaciones', 'Articulos', 'Responsables']
    
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook: Workbook = None
        self.stats = {
            'sedes_creadas': 0,
            'ubicaciones_creadas': 0,
            'articulos_creados': 0,
            'responsables_creados': 0,
            'items_eliminados': 0,
            'catalogos_eliminados': 0,
            'items_creados': 0,
        }
        self.errors: List[str] = []
        
        # Maps para convertir labels del Excel a valores de BD
        self.maps = {
            'tipo_ubicacion': {label.lower(): value for value, label in TipoUbicacion.choices},
            'categoria_articulo': {label.lower(): value for value, label in CategoriaArticulo.choices},
            'estado_fisico': {label.lower(): value for value, label in EstadoFisico.choices},
            'disponibilidad': {label.lower(): value for value, label in Disponibilidad.choices},
            'tipo_documento': {label.lower(): value for value, label in TipoDocumento.choices},
            'cargo_responsable': {label.lower(): value for value, label in CargoResponsable.choices},
        }
    
    @transaction.atomic
    def execute(self) -> Dict[str, Any]:
        try:
            self._load_workbook()
            self._validate_structure()
            
            sedes_data = self._parse_sedes_sheet()
            ubicaciones_data = self._parse_ubicaciones_sheet()
            articulos_data = self._parse_articulos_sheet()
            responsables_data = self._parse_responsables_sheet()
            items_data = self._parse_items_sheet()
            
            # Reseteo completo: Items y Catálogos
            self._delete_all_data()
            
            sedes_map = self._process_sedes(sedes_data)
            ubicaciones_map = self._process_ubicaciones(ubicaciones_data, sedes_map)
            articulos_map = self._process_articulos(articulos_data)
            responsables_map = self._process_responsables(responsables_data, sedes_map)
            
            self._process_items(
                items_data,
                sedes_map,
                ubicaciones_map,
                articulos_map,
                responsables_map
            )
            
            logger.info(f"Importación masiva completada: {self.stats}")
            
            return {
                'success': True,
                'stats': self.stats,
                'errors': self.errors
            }
            
        except Exception as e:
            self.errors.append(f"Error general: {str(e)}")
            raise ValidationError(f"Error en el proceso de importación: {str(e)}")
    
    def _load_workbook(self):
        try:
            self.workbook = load_workbook(self.excel_file, data_only=True)
        except Exception as e:
            raise ValidationError(f"Error al cargar el archivo Excel: {str(e)}")
    
    def _validate_structure(self):
        sheet_names = self.workbook.sheetnames
        for required_sheet in self.REQUIRED_SHEETS:
            if required_sheet not in sheet_names:
                raise ValidationError(f"Falta la hoja requerida: '{required_sheet}'")
                
    def _get_value_from_label(self, map_name: str, label: str) -> str:
        """
        Convierte el label (ej: 'Aula') al value (ej: 'aula').
        Si no encuentra match, retorna el valor original normalizado (para flexibilidad).
        """
        if not label:
            return ''
        
        label_clean = str(label).strip()
        label_lower = label_clean.lower()
        
        # Intentar buscar en el mapa de labels
        value = self.maps[map_name].get(label_lower)
        if value:
            return value
            
        # Si no encuentra, intentar ver si ya es el value (por compatibilidad)
        valid_values = self.maps[map_name].values()
        if label_lower in valid_values:
            return label_lower
            
        # Si no encuentra match, retornar el valor original (Flexibilidad total)
        # Esto permite valores personalizados como "Aula de Música" o "Regular-Malo"
        return label_clean

    def _parse_sedes_sheet(self) -> List[Dict[str, Any]]:
        sheet = self.workbook['Sedes']
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            
            # Validar obligatorios (Nombre*, Código*)
            if not row_dict.get('Nombre*'):
                self.errors.append(f"Sedes fila {row_idx}: 'Nombre' es obligatorio")
                continue
            if not row_dict.get('Código*'):
                self.errors.append(f"Sedes fila {row_idx}: 'Código' es obligatorio")
                continue
            
            data.append({
                'nombre': str(row_dict['Nombre*']).strip(),
                'codigo': str(row_dict['Código*']).strip().upper(),
                'direccion': str(row_dict.get('Dirección', '')).strip() if row_dict.get('Dirección') else '',
                'telefono': str(row_dict.get('Teléfono', '')).strip() if row_dict.get('Teléfono') else '',
                'email': str(row_dict.get('Email', '')).strip() if row_dict.get('Email') else '',
            })
        return data
    
    def _parse_ubicaciones_sheet(self) -> List[Dict[str, Any]]:
        sheet = self.workbook['Ubicaciones']
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            
            if not row_dict.get('Sede (Nombre)*'):
                self.errors.append(f"Ubicaciones fila {row_idx}: 'Sede (Nombre)' es obligatorio")
                continue
            if not row_dict.get('Nombre*'):
                self.errors.append(f"Ubicaciones fila {row_idx}: 'Nombre' es obligatorio")
                continue
            if not row_dict.get('Código*'):
                self.errors.append(f"Ubicaciones fila {row_idx}: 'Código' es obligatorio")
                continue
            if not row_dict.get('Tipo*'):
                self.errors.append(f"Ubicaciones fila {row_idx}: 'Tipo' es obligatorio")
                continue
            
            # Mapear Label -> Value (o usar custom)
            tipo_label = str(row_dict['Tipo*']).strip()
            tipo = self._get_value_from_label('tipo_ubicacion', tipo_label)
            
            data.append({
                'sede_nombre': str(row_dict['Sede (Nombre)*']).strip(),
                'nombre': str(row_dict['Nombre*']).strip(),
                'codigo': str(row_dict['Código*']).strip().upper(),
                'tipo': tipo,
                'piso': int(row_dict['Piso']) if row_dict.get('Piso') else None,
                'capacidad': int(row_dict['Capacidad']) if row_dict.get('Capacidad') else None,
                'observaciones': str(row_dict.get('Observaciones', '')).strip() if row_dict.get('Observaciones') else '',
            })
        return data
    
    def _parse_articulos_sheet(self) -> List[Dict[str, Any]]:
        sheet = self.workbook['Articulos']
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            
            if not row_dict.get('Nombre*'):
                self.errors.append(f"Articulos fila {row_idx}: 'Nombre' es obligatorio")
                continue
            if not row_dict.get('Categoría*'):
                self.errors.append(f"Articulos fila {row_idx}: 'Categoría' es obligatorio")
                continue
            
            # Mapear Label -> Value (o usar custom)
            cat_label = str(row_dict['Categoría*']).strip()
            categoria = self._get_value_from_label('categoria_articulo', cat_label)
            
            data.append({
                'nombre': str(row_dict['Nombre*']).strip(),
                'categoria': categoria,
                'codigo': str(row_dict.get('Código', '')).strip().upper() if row_dict.get('Código') else '',
                'descripcion': str(row_dict.get('Descripción', '')).strip() if row_dict.get('Descripción') else '',
            })
        return data
    
    def _parse_responsables_sheet(self) -> List[Dict[str, Any]]:
        sheet = self.workbook['Responsables']
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            
            # Validar nombre completo (unificado)
            nombre_completo = row_dict.get('Nombre Completo*')
            if not nombre_completo:
                self.errors.append(f"Responsables fila {row_idx}: 'Nombre Completo' es obligatorio")
                continue
            
            # Lógica de separación Nombre/Apellido
            partes = str(nombre_completo).strip().split()
            if len(partes) == 1:
                nombre = partes[0]
                apellido = '.' # Apellido dummy para cumplir restricción de modelo
            else:
                nombre = partes[0]
                apellido = ' '.join(partes[1:])
            
            # Validar Enums (o custom)
            tipo_doc = ''
            if row_dict.get('Tipo Documento'):
                label = str(row_dict['Tipo Documento']).strip()
                tipo_doc = self._get_value_from_label('tipo_documento', label)
            
            cargo = ''
            if row_dict.get('Cargo'):
                label = str(row_dict['Cargo']).strip()
                cargo = self._get_value_from_label('cargo_responsable', label)
            
            data.append({
                'nombre': nombre,
                'apellido': apellido,
                'nombre_completo': str(nombre_completo).strip(), # Para el mapa
                'tipo_documento': tipo_doc,
                'documento': str(row_dict.get('Documento', '')).strip() if row_dict.get('Documento') else '',
                'cargo': cargo,
                'email': str(row_dict.get('Email', '')).strip() if row_dict.get('Email') else '',
                'telefono': str(row_dict.get('Teléfono', '')).strip() if row_dict.get('Teléfono') else '',
                'sede_nombre': str(row_dict.get('Sede (Nombre)', '')).strip() if row_dict.get('Sede (Nombre)') else '',
            })
        return data
    
    def _parse_items_sheet(self) -> List[Dict[str, Any]]:
        sheet = self.workbook['Items']
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            
            # Validar obligatorios
            if not row_dict.get('Sede (Nombre)*'):
                self.errors.append(f"Items fila {row_idx}: 'Sede (Nombre)' es obligatorio")
                continue
            if not row_dict.get('Ubicacion (Nombre)*'):
                self.errors.append(f"Items fila {row_idx}: 'Ubicacion (Nombre)' es obligatorio")
                continue
            if not row_dict.get('Articulo (Nombre)*'):
                self.errors.append(f"Items fila {row_idx}: 'Articulo (Nombre)' es obligatorio")
                continue
                
            # Validar Enums (o custom)
            estado_label = str(row_dict.get('Estado Físico*', 'Bueno')).strip()
            estado = self._get_value_from_label('estado_fisico', estado_label)
                
            disp_label = str(row_dict.get('Disponibilidad*', 'En uso')).strip()
            disponibilidad = self._get_value_from_label('disponibilidad', disp_label)
            
            data.append({
                'sede_nombre': str(row_dict['Sede (Nombre)*']).strip(),
                'ubicacion_nombre': str(row_dict['Ubicacion (Nombre)*']).strip(),
                'articulo_nombre': str(row_dict['Articulo (Nombre)*']).strip(),
                'responsable_nombre': str(row_dict.get('Responsable (Nombre Completo)', '')).strip() if row_dict.get('Responsable (Nombre Completo)') else '',
                'placa': str(row_dict.get('Placa', '')).strip() if row_dict.get('Placa') else '',
                'marca': str(row_dict.get('Marca', '')).strip() if row_dict.get('Marca') else '',
                'serial': str(row_dict.get('Serial', '')).strip() if row_dict.get('Serial') else '',
                'estado': estado,
                'disponibilidad': disponibilidad,
                'descripcion': str(row_dict.get('Descripción', '')).strip() if row_dict.get('Descripción') else '',
                'observaciones': str(row_dict.get('Observaciones', '')).strip() if row_dict.get('Observaciones') else '',
            })
        return data

    # --- Métodos de procesamiento (BD) ---
    
    def _delete_all_data(self):
        """
        Elimina TODOS los datos del inventario (Items y Catálogos).
        Debe ejecutarse en orden estricto por Foreign Keys.
        """
        items_count = ItemInventario.objects.count()
        ItemInventario.objects.all().delete()
        self.stats['items_eliminados'] = items_count
        
        Ubicacion.objects.all().delete()
        Responsable.objects.all().delete()
        Sede.objects.all().delete()
        Articulo.objects.all().delete()
        
        self.stats['catalogos_eliminados'] = 1
        
    def _process_sedes(self, sedes_data: List[Dict]) -> Dict[str, Sede]:
        sedes_map = {}
        for d in sedes_data:
            try:
                sede, created = Sede.objects.update_or_create(
                    nombre=d['nombre'],
                    defaults={
                        'codigo': d['codigo'],
                        'direccion': d['direccion'],
                        'telefono': d['telefono'],
                        'email': d['email'],
                        'activo': True,
                    }
                )
                sedes_map[sede.nombre] = sede
                if created: self.stats['sedes_creadas'] += 1
            except Exception as e:
                self.errors.append(f"Error sede '{d['nombre']}': {str(e)}")
        return sedes_map

    def _process_ubicaciones(self, data: List[Dict], sedes_map: Dict) -> Dict:
        ubic_map = {}
        for d in data:
            sede = sedes_map.get(d['sede_nombre'])
            if not sede:
                self.errors.append(f"Ubicación '{d['nombre']}': Sede '{d['sede_nombre']}' no existe")
                continue
            try:
                ubic, created = Ubicacion.objects.update_or_create(
                    sede=sede,
                    codigo=d['codigo'],
                    defaults={
                        'nombre': d['nombre'],
                        'tipo': d['tipo'],
                        'piso': d['piso'],
                        'capacidad': d['capacidad'],
                        'observaciones': d['observaciones'],
                        'activo': True,
                    }
                )
                ubic_map[(sede.nombre, ubic.nombre)] = ubic
                if created: self.stats['ubicaciones_creadas'] += 1
            except Exception as e:
                self.errors.append(f"Error ubicación '{d['nombre']}': {str(e)}")
        return ubic_map

    def _process_articulos(self, data: List[Dict]) -> Dict:
        art_map = {}
        for d in data:
            try:
                if d['codigo']:
                    art, created = Articulo.objects.update_or_create(
                        codigo=d['codigo'],
                        defaults={
                            'nombre': d['nombre'],
                            'categoria': d['categoria'],
                            'descripcion': d['descripcion'],
                            'activo': True
                        }
                    )
                else:
                    art, created = Articulo.objects.update_or_create(
                        nombre=d['nombre'],
                        defaults={
                            'categoria': d['categoria'],
                            'descripcion': d['descripcion'],
                            'activo': True
                        }
                    )
                art_map[art.nombre] = art
                if created: self.stats['articulos_creados'] += 1
            except Exception as e:
                self.errors.append(f"Error artículo '{d['nombre']}': {str(e)}")
        return art_map

    def _process_responsables(self, data: List[Dict], sedes_map: Dict) -> Dict:
        resp_map = {}
        for d in data:
            sede = sedes_map.get(d['sede_nombre']) if d['sede_nombre'] else None
            
            try:
                defaults = {
                    'tipo_documento': d['tipo_documento'] if d['tipo_documento'] else None,
                    'documento': d['documento'] if d['documento'] else None,
                    'cargo': d['cargo'],
                    'email': d['email'] if d['email'] else None,
                    'telefono': d['telefono'],
                    'activo': True,
                }
                if sede:
                    defaults['sede'] = sede

                # Estrategia de deduplicación robusta (Email -> Documento -> Nombre)
                resp = None

                # 1. Buscar por Email (único globalmente)
                if d['email']:
                    resp = Responsable.objects.filter(email=d['email']).first()
                
                # 2. Buscar por Documento (único globalmente junto con tipo)
                if not resp and d['documento'] and d['tipo_documento']:
                    resp = Responsable.objects.filter(
                        documento=d['documento'], 
                        tipo_documento=d['tipo_documento']
                    ).first()

                # 3. Buscar por Nombre + Apellido (dentro de la misma sede o sin sede)
                if not resp:
                    query = Responsable.objects.filter(
                        nombre=d['nombre'], 
                        apellido=d['apellido']
                    )
                    if sede:
                        query = query.filter(sede=sede)
                    else:
                        query = query.filter(sede__isnull=True)
                    resp = query.first()

                if resp:
                    # Actualizar existente (último gana)
                    for key, value in defaults.items():
                        setattr(resp, key, value)
                    # También actualizar nombre/apellido por si acaso cambió levemente
                    resp.nombre = d['nombre']
                    resp.apellido = d['apellido']
                    resp.save()
                    created = False
                else:
                    # Crear nuevo
                    create_kwargs = {
                        'nombre': d['nombre'],
                        'apellido': d['apellido'],
                        **defaults
                    }
                    if 'sede' not in create_kwargs:
                        create_kwargs['sede'] = None
                        
                    resp = Responsable.objects.create(**create_kwargs)
                    created = True

                key = (d['nombre_completo'], d['sede_nombre'] or '')
                resp_map[key] = resp
                if created: self.stats['responsables_creados'] += 1

            except Exception as e:
                self.errors.append(f"Error responsable '{d['nombre_completo']}': {str(e)}")
        return resp_map

    def _process_items(self, data: List[Dict], sedes_map, ubic_map, art_map, resp_map):
        for d in data:
            try:
                # 1. Sede (Lazy creation)
                sede_nombre = d['sede_nombre']
                sede = sedes_map.get(sede_nombre)
                if not sede:
                    # Auto-create Sede
                    codigo_sede = slugify(sede_nombre).upper()[:20]
                    sede, created = Sede.objects.get_or_create(
                        nombre=sede_nombre,
                        defaults={
                            'codigo': codigo_sede,
                            'activo': True
                        }
                    )
                    sedes_map[sede_nombre] = sede
                    if created:
                        self.stats['sedes_creadas'] += 1

                # 2. Ubicación (Lazy creation)
                ubic_nombre = d['ubicacion_nombre']
                ubic_key = (sede_nombre, ubic_nombre)
                ubic = ubic_map.get(ubic_key)
                if not ubic:
                    # Auto-create Ubicacion
                    codigo_ubic = slugify(ubic_nombre).upper()[:50]
                    ubic, created = Ubicacion.objects.get_or_create(
                        sede=sede,
                        nombre=ubic_nombre,
                        defaults={
                            'codigo': codigo_ubic,
                            'tipo': 'otro', # Default for lazy creation (items sheet doesn't have type)
                            'activo': True
                        }
                    )
                    ubic_map[ubic_key] = ubic
                    if created:
                        self.stats['ubicaciones_creadas'] += 1

                # 3. Artículo (Lazy creation)
                art_nombre = d['articulo_nombre']
                art = art_map.get(art_nombre)
                if not art:
                    # Auto-create Articulo
                    art, created = Articulo.objects.get_or_create(
                        nombre=art_nombre,
                        defaults={
                            'categoria': 'otros', # Default
                            'activo': True
                        }
                    )
                    art_map[art_nombre] = art
                    if created:
                        self.stats['articulos_creados'] += 1
                
                # 4. Responsable (Lazy creation)
                resp_nombre_completo = d['responsable_nombre']
                resp = None
                if resp_nombre_completo:
                    # Try strict match first
                    resp = resp_map.get((resp_nombre_completo, sede_nombre))
                    if not resp:
                        resp = resp_map.get((resp_nombre_completo, ''))
                    
                    if not resp:
                        # Auto-create Responsable
                        partes = resp_nombre_completo.strip().split()
                        if len(partes) == 1:
                            nombre = partes[0]
                            apellido = '.'
                        else:
                            nombre = partes[0]
                            apellido = ' '.join(partes[1:])
                            
                        resp, created = Responsable.objects.get_or_create(
                            nombre=nombre,
                            apellido=apellido,
                            sede=sede,
                            defaults={
                                'activo': True
                            }
                        )
                        # Add to map to avoid re-creating for same person in same batch
                        resp_map[(resp_nombre_completo, sede_nombre)] = resp
                        if created:
                            self.stats['responsables_creados'] += 1
                
                # Limpieza inteligente de seriales (User Feedback: Seriales genéricos causan duplicados)
                serial = d['serial']
                if serial:
                    serial_limpio = serial.upper().strip()
                    # Si es un placeholder común, convertir a None para permitir múltiples ítems
                    if any(x in serial_limpio for x in ['NO TIENE', 'NO SE VE', 'S/N', 'SIN SERIAL', 'NO APLICA', 'N/A', 'NINGUNO']):
                        serial = None
                    # Si parece ser un modelo en vez de un serial único (ej: "MODEL:PCS...")
                    elif serial_limpio.startswith('MODEL:'):
                        serial = None
                
                # Limpieza de Placa (para evitar error de unicidad con strings vacíos)
                placa = d['placa']
                if placa:
                    placa_limpia = placa.upper().strip()
                    if not placa_limpia or any(x in placa_limpia for x in ['NO TIENE', 'NO APLICA', 'S/P', 'SIN PLACA', 'N/A', 'NINGUNA']):
                        placa = None
                    else:
                        placa = placa_limpia # Usar la versión limpia
                else:
                    placa = None

                # Crear siempre un nuevo registro (ya no hay restricción de serial único)
                ItemInventario.objects.create(
                    articulo=art,
                    ubicacion=ubic,
                    sede=sede,
                    responsable=resp,
                    placa=placa,
                    marca=d['marca'],
                    serial=serial,
                    estado=d['estado'],
                    disponibilidad=d['disponibilidad'],
                    descripcion=d['descripcion'],
                    observaciones=d['observaciones']
                )
                self.stats['items_creados'] += 1

            except Exception as e:
                self.errors.append(f"Error creando ítem: {str(e)}")
