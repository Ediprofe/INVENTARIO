"""
Generador de archivo Excel para exportación completa (compatible con importación).
"""

import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils.encoding import force_str

# Evitar importaciones globales de modelos para prevenir dependencias circulares
from apps.inventario.models.choices import (
    TipoUbicacion,
    CategoriaArticulo,
    EstadoFisico,
    Disponibilidad,
    TipoDocumento,
    CargoResponsable,
)

def generate_full_export_workbook() -> Workbook:
    """
    Genera un workbook con todas las hojas y datos del sistema,
    en el formato exacto requerido para la importación/reseteo.
    """
    # Importación diferida de modelos
    from apps.inventario.models import Sede, Ubicacion, Articulo, Responsable, ItemInventario

    wb = Workbook()
    
    try:
        # Crear las 5 hojas en orden
        _create_items_sheet_with_data(wb, ItemInventario)
        _create_sedes_sheet_with_data(wb, Sede)
        _create_ubicaciones_sheet_with_data(wb, Ubicacion)
        _create_articulos_sheet_with_data(wb, Articulo)
        _create_responsables_sheet_with_data(wb, Responsable)
        
        # Eliminar hoja por defecto si existe
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Activar la hoja 'Items'
        if 'Items' in wb.sheetnames:
            try:
                wb.active = wb.sheetnames.index('Items')
            except ValueError:
                wb.active = 0
                
    except Exception as e:
        print(f"Error generando workbook: {str(e)}")
        traceback.print_exc()
        raise e
    
    return wb

def _sanitize_value(value):
    """Limpia valores para asegurar compatibilidad con Excel."""
    if value is None:
        return ""
    # Convertir a string usando force_str de Django para manejar lazy translations y objetos proxied
    val_str = force_str(value)
    # Remover caracteres de control ilegales en Excel (excepto tab, CR, LF)
    # Excel no soporta caracteres ASCII 0-8, 11-12, 14-31
    return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', val_str)

def _get_choice_label(value, choices_class):
    """Obtiene el label legible de una opción."""
    if not value:
        return ""
    for choice in choices_class.choices:
        if choice[0] == value:
            return _sanitize_value(choice[1])
    return _sanitize_value(value)

def _apply_header_style(ws, headers: list):
    """Aplica estilos a la fila de encabezados."""
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 22

def _auto_adjust_columns(ws, headers: list, max_width: int = 50):
    """Ajusta automáticamente el ancho de las columnas basado en el contenido."""
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))
        
        # Revisar todas las celdas de la columna para encontrar el contenido más largo
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    # Contar líneas si hay saltos de línea
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
        
        # Ajustar ancho con un margen adicional, limitado al máximo
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def _apply_wrap_text_to_cells(ws):
    """Aplica wrap_text a todas las celdas de datos (excluyendo encabezados)."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

def _convert_to_table(ws, table_name: str, headers: list, data_row_count: int):
    """Convierte un rango de celdas en una tabla de Excel con formato."""
    if data_row_count == 0:
        # Si no hay datos, solo crear tabla con encabezados
        end_row = 1
    else:
        end_row = data_row_count + 1  # +1 porque la fila 1 es encabezado
    
    end_column = get_column_letter(len(headers))
    table_ref = f"A1:{end_column}{end_row}"
    
    # Crear tabla
    table = Table(displayName=table_name, ref=table_ref)
    
    # Aplicar estilo de tabla
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Estilo azul similar al encabezado actual
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Agregar tabla a la hoja
    ws.add_table(table)

def _add_data_validation(ws, column_letter: str, options: list, allow_blank: bool = True):
    """Agrega validación de datos a una columna."""
    try:
        # Limpiar opciones
        clean_options = [_sanitize_value(opt).replace(',', ';') for opt in options] # Reemplazar comas internas
        
        # Excel tiene límite de 255 caracteres para listas explícitas
        formula = f'"{",".join(clean_options)}"'
        
        if len(formula) > 255:
            return 
            
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank,
            showDropDown=True
        )
        dv.add(f'{column_letter}2:{column_letter}5000')
        ws.add_data_validation(dv)
    except Exception:
        pass

def _create_sedes_sheet_with_data(wb: Workbook, SedeModel):
    ws = wb.create_sheet('Sedes')
    headers = ['Nombre*', 'Código*', 'Coordinador', 'Dirección', 'Teléfono', 'Email']
    _apply_header_style(ws, headers)
    
    sedes = SedeModel.objects.select_related('coordinador').all().order_by('nombre')
    row_count = 0
    for row_num, obj in enumerate(sedes, 2):
        ws.cell(row=row_num, column=1, value=_sanitize_value(obj.nombre))
        ws.cell(row=row_num, column=2, value=_sanitize_value(obj.codigo))
        
        coordinador_nombre = obj.coordinador.nombre_completo if obj.coordinador else ""
        ws.cell(row=row_num, column=3, value=_sanitize_value(coordinador_nombre))
        
        ws.cell(row=row_num, column=4, value=_sanitize_value(obj.direccion))
        ws.cell(row=row_num, column=5, value=_sanitize_value(obj.telefono))
        ws.cell(row=row_num, column=6, value=_sanitize_value(obj.email))
        row_count += 1
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaSedes', headers, row_count)

def _create_ubicaciones_sheet_with_data(wb: Workbook, UbicacionModel):
    ws = wb.create_sheet('Ubicaciones')
    headers = ['Sede (Nombre)*', 'Nombre*', 'Código*', 'Tipo*', 'Responsable Por Defecto', 'Piso', 'Capacidad', 'Observaciones']
    _apply_header_style(ws, headers)
    
    # Validación Tipo
    tipos_labels = [_sanitize_value(choice[1]) for choice in TipoUbicacion.choices]
    _add_data_validation(ws, 'D', tipos_labels, allow_blank=False)
    
    ubicaciones = UbicacionModel.objects.select_related('sede', 'responsable').all().order_by('sede__nombre', 'nombre')
    row_count = 0
    for row_num, obj in enumerate(ubicaciones, 2):
        ws.cell(row=row_num, column=1, value=_sanitize_value(obj.sede.nombre))
        ws.cell(row=row_num, column=2, value=_sanitize_value(obj.nombre))
        ws.cell(row=row_num, column=3, value=_sanitize_value(obj.codigo))
        ws.cell(row=row_num, column=4, value=_sanitize_value(obj.tipo))
        
        responsable_nombre = obj.responsable.nombre_completo if obj.responsable else ""
        ws.cell(row=row_num, column=5, value=_sanitize_value(responsable_nombre))
        
        ws.cell(row=row_num, column=6, value=obj.piso if obj.piso is not None else "")
        ws.cell(row=row_num, column=7, value=obj.capacidad if obj.capacidad is not None else "")
        ws.cell(row=row_num, column=8, value=_sanitize_value(obj.observaciones))
        row_count += 1
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaUbicaciones', headers, row_count)

def _create_articulos_sheet_with_data(wb: Workbook, ArticuloModel):
    ws = wb.create_sheet('Articulos')
    headers = ['Nombre*', 'Categoría*', 'Código', 'Descripción']
    _apply_header_style(ws, headers)
    
    cat_labels = [_sanitize_value(choice[1]) for choice in CategoriaArticulo.choices]
    _add_data_validation(ws, 'B', cat_labels, allow_blank=False)
    
    articulos = ArticuloModel.objects.all().order_by('nombre')
    row_count = 0
    for row_num, obj in enumerate(articulos, 2):
        ws.cell(row=row_num, column=1, value=_sanitize_value(obj.nombre))
        cat_label = _get_choice_label(obj.categoria, CategoriaArticulo)
        ws.cell(row=row_num, column=2, value=cat_label)
        # Usar atributo 'codigo' que es el estándar del modelo Articulo
        # Si en el futuro se añade 'codigo_bien', se puede usar getattr con default
        codigo_val = getattr(obj, 'codigo', getattr(obj, 'codigo_bien', ''))
        ws.cell(row=row_num, column=3, value=_sanitize_value(codigo_val))
        ws.cell(row=row_num, column=4, value=_sanitize_value(obj.descripcion))
        row_count += 1
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaArticulos', headers, row_count)

def _create_responsables_sheet_with_data(wb: Workbook, ResponsableModel):
    ws = wb.create_sheet('Responsables')
    headers = [
        'Nombre Completo*', 'Tipo Documento', 'Documento', 'Cargo', 
        'Email', 'Teléfono', 'Sede (Nombre)'
    ]
    _apply_header_style(ws, headers)
    
    # Validaciones
    tipo_doc_labels = [_sanitize_value(choice[1]) for choice in TipoDocumento.choices]
    _add_data_validation(ws, 'B', tipo_doc_labels)
    cargo_labels = [_sanitize_value(choice[1]) for choice in CargoResponsable.choices]
    _add_data_validation(ws, 'D', cargo_labels)
    
    responsables = ResponsableModel.objects.select_related('sede').all().order_by('apellido', 'nombre')
    row_count = 0
    for row_num, obj in enumerate(responsables, 2):
        ws.cell(row=row_num, column=1, value=_sanitize_value(obj.nombre_completo))
        
        td_label = _get_choice_label(obj.tipo_documento, TipoDocumento)
        ws.cell(row=row_num, column=2, value=td_label)
        
        ws.cell(row=row_num, column=3, value=_sanitize_value(obj.documento))
        
        cargo_label = _get_choice_label(obj.cargo, CargoResponsable)
        ws.cell(row=row_num, column=4, value=cargo_label)
        
        ws.cell(row=row_num, column=5, value=_sanitize_value(obj.email))
        ws.cell(row=row_num, column=6, value=_sanitize_value(obj.telefono))
        
        sede_nombre = obj.sede.nombre if obj.sede else ""
        ws.cell(row=row_num, column=7, value=_sanitize_value(sede_nombre))
        row_count += 1
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaResponsables', headers, row_count)

def _create_items_sheet_with_data(wb: Workbook, ItemModel):
    ws = wb.create_sheet('Items')
    headers = [
        'Sede (Nombre)*', 'Ubicacion (Nombre)*', 'Articulo (Nombre)*', 
        'Responsable (Nombre Completo)', 'Placa', 'Marca', 'Serial', 
        'Estado Físico*', 'Disponibilidad*', 'Descripción', 'Observaciones'
    ]
    _apply_header_style(ws, headers)
    
    # Validaciones
    estado_labels = [_sanitize_value(choice[1]) for choice in EstadoFisico.choices]
    _add_data_validation(ws, 'H', estado_labels, allow_blank=False)
    
    disp_labels = [_sanitize_value(choice[1]) for choice in Disponibilidad.choices]
    _add_data_validation(ws, 'I', disp_labels, allow_blank=False)
    
    # Query optimizada
    items = ItemModel.objects.select_related(
        'sede', 'ubicacion', 'articulo', 'responsable'
    ).all().order_by('-created_at')
    
    row_count = 0
    for row_num, obj in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=_sanitize_value(obj.sede.nombre))
        ws.cell(row=row_num, column=2, value=_sanitize_value(obj.ubicacion.nombre))
        ws.cell(row=row_num, column=3, value=_sanitize_value(obj.articulo.nombre))
        
        resp_nombre = obj.responsable.nombre_completo if obj.responsable else ""
        ws.cell(row=row_num, column=4, value=_sanitize_value(resp_nombre))
        
        ws.cell(row=row_num, column=5, value=_sanitize_value(obj.placa))
        ws.cell(row=row_num, column=6, value=_sanitize_value(obj.marca))
        ws.cell(row=row_num, column=7, value=_sanitize_value(obj.serial))
        
        ws.cell(row=row_num, column=8, value=_get_choice_label(obj.estado, EstadoFisico))
        ws.cell(row=row_num, column=9, value=_get_choice_label(obj.disponibilidad, Disponibilidad))
        
        ws.cell(row=row_num, column=10, value=_sanitize_value(obj.descripcion))
        ws.cell(row=row_num, column=11, value=_sanitize_value(obj.observaciones))
        row_count += 1
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaItems', headers, row_count)
