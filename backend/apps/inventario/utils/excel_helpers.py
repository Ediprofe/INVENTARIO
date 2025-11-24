"""
Funciones auxiliares para generación de archivos Excel.

Este módulo contiene funciones de utilidad para crear y formatear
archivos Excel de exportación.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def apply_header_styles(ws, headers: list[str]) -> None:
    """
    Aplica estilos al encabezado de una hoja Excel.
    
    Args:
        ws: Worksheet de openpyxl
        headers: Lista de nombres de columnas
    """
    header_fill = PatternFill(
        start_color="0066CC", 
        end_color="0066CC", 
        fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def adjust_column_widths(ws, max_width: int = 50) -> None:
    """
    Ajusta automáticamente el ancho de las columnas.
    
    Args:
        ws: Worksheet de openpyxl
        max_width: Ancho máximo de columna en caracteres
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def write_item_row(ws, row_num: int, item) -> None:
    """
    Escribe los datos de un ítem en una fila del Excel.
    
    Args:
        ws: Worksheet de openpyxl
        row_num: Número de fila donde escribir
        item: Instancia de ItemInventario
    """
    ws.cell(row=row_num, column=1, value=item.id)
    ws.cell(row=row_num, column=2, value=item.codigo)
    ws.cell(row=row_num, column=3, value=item.placa or '')
    ws.cell(row=row_num, column=4, value=item.articulo.nombre)
    ws.cell(row=row_num, column=5, value=item.articulo.codigo)
    ws.cell(row=row_num, column=6, value=item.articulo.get_categoria_display())
    ws.cell(row=row_num, column=7, value=item.sede.nombre)
    ws.cell(row=row_num, column=8, value=item.ubicacion.nombre)
    ws.cell(row=row_num, column=9, value=item.ubicacion.codigo)
    ws.cell(row=row_num, column=10, value=item.responsable.nombre_completo if item.responsable else '')
    ws.cell(row=row_num, column=11, value=item.responsable.documento if item.responsable else '')
    ws.cell(row=row_num, column=12, value=item.marca or '')
    ws.cell(row=row_num, column=13, value=item.serial or '')
    ws.cell(row=row_num, column=14, value=item.estado)
    ws.cell(row=row_num, column=15, value=item.disponibilidad)
    ws.cell(row=row_num, column=16, value=item.descripcion or '')
    ws.cell(row=row_num, column=17, value=item.observaciones or '')
    ws.cell(row=row_num, column=18, value=item.created_at.strftime('%Y-%m-%d %H:%M'))
    ws.cell(row=row_num, column=19, value=item.updated_at.strftime('%Y-%m-%d %H:%M'))


def write_template_example_row(ws) -> None:
    """
    Escribe una fila de ejemplo en la plantilla de importación.
    
    Args:
        ws: Worksheet de openpyxl
    """
    ws.cell(row=2, column=1, value='Sede Central')
    ws.cell(row=2, column=2, value='A-101')
    ws.cell(row=2, column=3, value='Portátil Lenovo')
    ws.cell(row=2, column=4, value='Bueno')
    ws.cell(row=2, column=5, value='En uso')
    ws.cell(row=2, column=6, value='Juan Pérez')
    ws.cell(row=2, column=7, value='PLA-001')
    ws.cell(row=2, column=8, value='Lenovo')
    ws.cell(row=2, column=9, value='SN123456')
    ws.cell(row=2, column=10, value='Descripción del ítem')
    ws.cell(row=2, column=11, value='Observaciones adicionales')


# Constantes de columnas
EXPORT_HEADERS = [
    'ID', 'Código Ítem', 'Placa', 'Artículo', 'Artículo Código', 
    'Categoría', 'Sede', 'Ubicación', 'Ubicación Código', 'Responsable',
    'Responsable Documento', 'Marca', 'Serial', 'Estado Físico', 
    'Disponibilidad', 'Descripción', 'Observaciones', 'Creado', 
    'Actualizado'
]

IMPORT_TEMPLATE_HEADERS = [
    'Sede', 'Ubicacion', 'Articulo', 'Estado', 'Disponibilidad', 
    'Responsable', 'Placa', 'Marca', 'Serial', 'Descripcion', 
    'Observaciones'
]

