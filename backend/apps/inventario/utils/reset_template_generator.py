"""
Generador de plantilla Excel multi-hoja para reseteo e importación masiva.

Este módulo crea un archivo Excel con 5 hojas pre-llenadas con datos de ejemplo
que sirven como guía para el usuario. Incluye validaciones de datos (listas desplegables)
para guiar al usuario con los valores permitidos por el sistema.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from apps.inventario.models.choices import (
    TipoUbicacion,
    CategoriaArticulo,
    EstadoFisico,
    Disponibilidad,
    TipoDocumento,
    CargoResponsable,
)


def generate_reset_template() -> Workbook:
    """
    Genera plantilla Excel multi-hoja con datos de ejemplo y validaciones.
    
    La plantilla incluye 5 hojas:
    1. Items: Inventario principal
    2. Sedes: Catálogo de sedes
    3. Ubicaciones: Catálogo de ubicaciones
    4. Articulos: Catálogo de artículos
    5. Responsables: Catálogo de responsables
    
    Returns:
        Workbook de openpyxl listo para guardar
    """
    wb = Workbook()
    
    # Crear las 5 hojas
    _create_sedes_sheet(wb)
    _create_ubicaciones_sheet(wb)
    _create_articulos_sheet(wb)
    _create_responsables_sheet(wb)
    _create_items_sheet(wb)
    
    # Eliminar hoja por defecto si existe
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Activar la primera hoja (Items)
    wb.active = wb['Items']
    
    return wb


def _apply_header_style(ws, headers: list):
    """
    Aplica estilos a la fila de encabezados.
    Los campos obligatorios deben terminar en *.
    
    Args:
        ws: Worksheet de openpyxl
        headers: Lista de nombres de columnas
    """
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 22


def _add_data_validation(ws, column_letter: str, options: list, allow_blank: bool = True):
    """
    Agrega validación de datos (lista desplegable) a una columna completa.
    
    Args:
        ws: Worksheet
        column_letter: Letra de la columna (ej: 'D')
        options: Lista de opciones válidas
        allow_blank: Si se permite dejar vacío
    """
    # Convertir lista de opciones a string separado por comas
    # Nota: Excel tiene límite de 255 caracteres para la fórmula de lista explícita.
    # Si las opciones son muy largas, idealmente se deberían poner en una hoja oculta,
    # pero para este caso usaremos lista explícita si es corta.
    
    formula = f'"{",".join(options)}"'
    
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showDropDown=True,
        errorTitle="Entrada inválida",
        error="Por favor seleccione un valor de la lista."
    )
    
    # Aplicar a toda la columna (desde fila 2 hasta 1000 por ejemplo)
    dv.add(f'{column_letter}2:{column_letter}1000')
    ws.add_data_validation(dv)


def _auto_adjust_columns(ws, headers: list, max_width: int = 50):
    """Ajusta automáticamente el ancho de las columnas basado en el contenido."""
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))
        
        # Revisar todas las celdas de la columna para encontrar el contenido más largo
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    # Contar líneas si hay saltos de línea
                    cell_value = str(cell.value)
                    lines = cell_value.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
        
        # Ajustar ancho con un margen adicional, limitado al máximo
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _apply_wrap_text_to_cells(ws):
    """Aplica wrap_text a todas las celdas de datos (excluyendo encabezados)."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def _convert_to_table(ws, table_name: str, headers: list, data_row_count: int):
    """Convierte un rango de celdas en una tabla de Excel con formato."""
    if data_row_count == 0:
        # Si no hay datos, solo crear tabla con encabezados
        end_row = 1
    else:
        end_row = data_row_count + 1  # +1 porque la fila 1 es encabezado
    
    end_column = get_column_letter(len(headers))
    table_ref = f"A1:{end_column}{end_row}"
    
    # Crear tabla
    table = Table(displayName=table_name, ref=table_ref)
    
    # Aplicar estilo de tabla
    style = TableStyleInfo(
        name="TableStyleMedium9",  # Estilo azul similar al encabezado actual
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Agregar tabla a la hoja
    ws.add_table(table)


def _create_sedes_sheet(wb: Workbook):
    """
    Crea la hoja 'Sedes' con datos de ejemplo.
    """
    ws = wb.create_sheet('Sedes')
    
    # Campos obligatorios marcados con *
    headers = ['Nombre*', 'Código*', 'Dirección', 'Teléfono', 'Email']
    _apply_header_style(ws, headers)
    
    example_data = [
        ['Sede Central', 'CENTRAL', 'Calle 123 #45-67', '3001234567', 'central@institucion.edu.co'],
        ['Sede Norte', 'NORTE', 'Carrera 89 #10-11', '3002345678', 'norte@institucion.edu.co'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaSedes', headers, len(example_data))


def _create_ubicaciones_sheet(wb: Workbook):
    """
    Crea la hoja 'Ubicaciones' con validación de tipos.
    """
    ws = wb.create_sheet('Ubicaciones')
    
    headers = ['Sede (Nombre)*', 'Nombre*', 'Código*', 'Tipo*', 'Piso', 'Capacidad', 'Observaciones']
    _apply_header_style(ws, headers)
    
    # Validación para Tipo
    tipos_labels = [choice[1] for choice in TipoUbicacion.choices]  # Usar labels (ej: 'Aula')
    _add_data_validation(ws, 'D', tipos_labels, allow_blank=False)
    
    example_data = [
        ['Sede Central', 'Aula 101', 'AULA-101', 'Aula', 1, 30, 'Matemáticas'],
        ['Sede Norte', 'Oficina Principal', 'OF-01', 'Oficina', 2, 5, 'Administración'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaUbicaciones', headers, len(example_data))


def _create_articulos_sheet(wb: Workbook):
    """
    Crea la hoja 'Articulos' con validación de categorías.
    """
    ws = wb.create_sheet('Articulos')
    
    headers = ['Nombre*', 'Categoría*', 'Código', 'Descripción']
    _apply_header_style(ws, headers)
    
    # Validación para Categoría
    cat_labels = [choice[1] for choice in CategoriaArticulo.choices]
    _add_data_validation(ws, 'B', cat_labels, allow_blank=False)
    
    example_data = [
        ['Computador Portátil', 'Tecnología', 'TEC-001', 'Laptop Core i5'],
        ['Silla Universitaria', 'Mobiliario', 'MOB-001', 'Silla plástica'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaArticulos', headers, len(example_data))


def _create_responsables_sheet(wb: Workbook):
    """
    Crea la hoja 'Responsables'.
    NOTA: Se unifica Nombre y Apellido en 'Nombre Completo' para facilitar copia desde Excel.
    """
    ws = wb.create_sheet('Responsables')
    
    # Cambio solicitado: Unificar nombre y apellido
    headers = [
        'Nombre Completo*', 
        'Tipo Documento', 
        'Documento', 
        'Cargo', 
        'Email', 
        'Teléfono', 
        'Sede (Nombre)'
    ]
    _apply_header_style(ws, headers)
    
    # Validación Tipo Documento (Col B)
    tipo_doc_labels = [choice[1] for choice in TipoDocumento.choices]
    _add_data_validation(ws, 'B', tipo_doc_labels)
    
    # Validación Cargo (Col D)
    cargo_labels = [choice[1] for choice in CargoResponsable.choices]
    _add_data_validation(ws, 'D', cargo_labels)
    
    example_data = [
        ['Juan Pérez González', 'Cédula de Ciudadanía', '1234567890', 'Docente', 'juan@edu.co', '3001112233', 'Sede Central'],
        ['María López', 'Cédula de Ciudadanía', '9876543210', 'Coordinador', 'maria@edu.co', '3004445566', 'Sede Norte'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaResponsables', headers, len(example_data))


def _create_items_sheet(wb: Workbook):
    """
    Crea la hoja 'Items' con validaciones.
    """
    ws = wb.create_sheet('Items')
    
    headers = [
        'Sede (Nombre)*',
        'Ubicacion (Nombre)*',
        'Articulo (Nombre)*',
        'Responsable (Nombre Completo)',
        'Placa',
        'Marca',
        'Serial',
        'Estado Físico*',
        'Disponibilidad*',
        'Descripción',
        'Observaciones'
    ]
    _apply_header_style(ws, headers)
    
    # Validación Estado Físico (Col H)
    estado_labels = [choice[1] for choice in EstadoFisico.choices]
    _add_data_validation(ws, 'H', estado_labels, allow_blank=False)
    
    # Validación Disponibilidad (Col I)
    disp_labels = [choice[1] for choice in Disponibilidad.choices]
    _add_data_validation(ws, 'I', disp_labels, allow_blank=False)
    
    example_data = [
        ['Sede Central', 'Aula 101', 'Computador Portátil', 'Juan Pérez González', 'PC-001', 'Dell', 'XYZ123', 'Bueno', 'En uso', '', ''],
        ['Sede Norte', 'Oficina Principal', 'Silla Universitaria', '', 'SILLA-01', 'Rimax', '', 'Regular', 'En uso', '', ''],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formato de tabla, wrap text y autoajuste
    _apply_wrap_text_to_cells(ws)
    _auto_adjust_columns(ws, headers)
    _convert_to_table(ws, 'TablaItems', headers, len(example_data))
